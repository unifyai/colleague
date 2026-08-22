"""Deterministic spreadsheet writing, and the reading the scorers do.

Writing: openpyxl is byte-deterministic once the two timestamps in
docProps are pinned, which `write_workbook` does. `selftest` regenerates
corpora and compares bytes, so if a future openpyxl starts leaking
nondeterminism the suite fails loudly rather than silently shipping
unreproducible fixtures.

Reading: `read_workbook` is how a scorer parses the artifact an arm
returned. It is deliberately forgiving about *container* (any zip openpyxl
accepts, any sheet count) and deliberately literal about *content*: cell
values come back exactly as typed — a string stays a string, a float stays
a float, a bool stays a bool — because whether an amount arrived as the
string "812.40" or the number 812.4 is precisely the kind of thing a track
scores. No coercion happens here; tolerance is the scorer's declared
business, not the reader's.
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

#: Pinned into docProps of every generated workbook; see pdfgen.CREATION_DATE.
CREATION_DATE = datetime(2026, 1, 5, 9, 0, 0, tzinfo=timezone.utc)


def write_workbook(path: str | Path, sheets: dict[str, list[list[Any]]]) -> None:
    """Write ``{sheet_name: rows}`` deterministically.

    Rows are lists of cell values; ``None`` leaves a gap. Sheet order is
    the dict's insertion order.
    """
    wb = openpyxl.Workbook()
    wb.properties.created = CREATION_DATE
    wb.properties.modified = CREATION_DATE
    default = wb.active
    for i, (name, rows) in enumerate(sheets.items()):
        ws = default if i == 0 else wb.create_sheet()
        ws.title = name
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    Path(path).write_bytes(_pin_zip_times(buf.getvalue()))


_CORE_STAMP = re.compile(
    rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
)


def _pin_zip_times(data: bytes) -> bytes:
    """Rewrite the xlsx zip with fixed entry timestamps and docProps dates.

    Pinning `wb.properties` before save is not enough twice over: openpyxl
    restamps `dcterms:modified` at save time, and every zip member's local
    header carries mtime seconds — so two saves a second apart differ.
    Entry order is kept exactly as openpyxl wrote it.
    """
    src = zipfile.ZipFile(io.BytesIO(data))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        for info in src.infolist():
            content = src.read(info.filename)
            if info.filename == "docProps/core.xml":
                content = _CORE_STAMP.sub(rb"\g<1>2026-01-05T09:00:00Z\g<2>", content)
            pinned = zipfile.ZipInfo(info.filename, date_time=(2026, 1, 5, 9, 0, 0))
            pinned.external_attr = info.external_attr
            pinned.compress_type = zipfile.ZIP_DEFLATED
            z.writestr(pinned, content)
    return out.getvalue()


def read_workbook(source: str | Path | bytes) -> dict[str, list[list[Any]]]:
    """Every sheet's cells, values only, trailing empty rows dropped.

    Raises whatever openpyxl raises on a file that is not a workbook —
    callers treat that as "the returned artifact is not parseable", which
    is a scored fact, not an infrastructure error.
    """
    handle = io.BytesIO(source) if isinstance(source, bytes) else str(source)
    wb = openpyxl.load_workbook(handle, read_only=True, data_only=True)
    try:
        out: dict[str, list[list[Any]]] = {}
        for ws in wb.worksheets:
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            while rows and all(v is None for v in rows[-1]):
                rows.pop()
            out[ws.title] = rows
        return out
    finally:
        wb.close()
